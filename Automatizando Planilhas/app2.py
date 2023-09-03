import openpyxl

#Carregando o arquivo
book= openpyxl.load_workbook('Planilha de Compras.xlsx')
# Selecionando uma pagina
frutas_page= book['Frutas']
# Imprimindo dados de cada linha

#(min_row=2,max_row=5) seleciona o começo e o final da leitura da linha

#Imprimi em forma de lista
#for cell in rows:
#   print(cell.value)

#Dessa maneira imprimi tudo junto
for rows in frutas_page.iter_rows(min_row=2,max_row=5):
    print(f'{rows[0].value,rows[1].value,rows[2].value}')


#Para modificar a informação basta realizar a seguinte linha de codigo
#for rows in furtas_page.iter_rows(min_row=2, max_row=5):
#for cell in rows:
#if cell.value =='Banana':
#cell.value = 'Morango' (Só vai mudar se a linha cima for verdadeira)
#para salvar é book.save('Planilha de Compras')
#Caso queira criar outro arquivo com a mudannça só renomea dentro do ()