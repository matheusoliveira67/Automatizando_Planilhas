import openpyxl

#criar uma planilha     
book = openpyxl.Workbook()
#como visualizar paginas existentes 
print(book.sheetnames)
#como criar uma pagina
book.create_sheet('Frutas')
#como selecionar uma pagina
frutas_page = book['Frutas']
frutas_page.append(['Frutas', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Maça', '4', 'R$6,90'])
frutas_page.append(['Melancia', '6', 'R$12,90'])
frutas_page.append(['Acerola', '2', 'R$3,90'])

#Salvar a planilha
book.save('Planilha de Compras.xlsx')